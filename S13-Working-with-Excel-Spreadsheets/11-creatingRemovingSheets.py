import openpyxl
wb = openpyxl.Workbook()
wb.sheetnames

wb.create_sheet()  # Add a new sheet.

wb.sheetnames

# Create a new sheet at index 0.
wb.create_sheet(index=0, title='First Sheet')

wb.sheetnames

wb.create_sheet(index=2, title='Middle Sheet')

wb.sheetnames

del wb['Middle Sheet']
del wb['Sheet1']
wb.sheetnames

# Writing Values to Cells
import openpyxl
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, world!'  # Edit the cell's value.
sheet['A1'].value