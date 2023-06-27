import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']  # Get a sheet from the workbook.
sheet['A1']  # Get a cell from the sheet.

sheet['A1'].value # Get the value from the cell.

c = sheet['B1']  # Get another cell from the sheet.
c.value

# Get the row, column, and value from the cell.
'Row %s, Column %s is %s' % (c.row, c.column, c.value)

'Cell %s is %s' % (c.coordinate, c.value)

sheet['C1'].value

sheet.cell(row=1, column=2)

sheet.cell(row=1, column=2).value

for i in range(1, 8, 2):  # Go through every other row:
    print(i, sheet.cell(row=i, column=2).value)

sheet.max_row  # Get the highest row number.
sheet.max_column  # Get the highest column number.