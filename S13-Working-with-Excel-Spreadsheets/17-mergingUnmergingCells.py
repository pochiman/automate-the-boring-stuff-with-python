import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')  # Merge all these cells.
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')  # Merge these two cells.
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')

import openpyxl
wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3')  # Split these cells up.
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')